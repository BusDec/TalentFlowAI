"""Import candidate profiles from the TalentBridge template.

Supports:
- .xlsx workbooks with all four sheets (CANDIDATE_BIO, ACADEMIC_RECORD,
  WORK_EXP, EXAM_DISCLOSURE_PORTAL). The candidate is identified by the Email
  in CANDIDATE_BIO; all rows in the other sheets belong to that candidate.
- .csv files in CANDIDATE_BIO format (one bio per row).
"""

import csv
import datetime
import io

from recruitment.models import Candidate

from .models import (
    AcademicRecord,
    CandidateProfile,
    ExamDisclosure,
    WorkExperience,
)

_CATEGORY_MAP = {
    "ur": "ur", "general": "ur", "obc": "obc", "obc (ncl)": "obc",
    "sc": "sc", "st": "st", "ews": "ews",
}
_GENDER_MAP = {"male": "M", "m": "M", "female": "F", "f": "F", "other": "O", "o": "O"}

_LEVEL_MAP = {
    "10th": "10th", "ssc": "10th", "ssc / 10th": "10th", "matric": "10th",
    "12th": "12th", "hsc": "12th", "hsc / 12th": "12th", "intermediate": "12th",
    "diploma": "diploma", "polytechnic": "diploma",
    "ug": "ug", "bachelor": "ug", "btech": "ug", "graduate": "ug",
    "pg": "pg", "master": "pg", "post-graduate": "pg", "post graduate": "pg",
}
_MARKING_MAP = {"percentage": "percentage", "%": "percentage", "cgpa": "cgpa", "grade": "grade"}
_EXAM_MAP = {"gate": "gate", "ese": "ese", "gate + ese": "both", "gate+ese": "both"}


def _parse_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    raw = str(raw).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _to_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", "")))
    except (TypeError, ValueError):
        return None


def _norm(value, mapping, default=""):
    return mapping.get(str(value or "").strip().lower(), default)


def _row_to_dict(header, row):
    return {str(header[i]).strip() if header[i] else f"C{i}": row[i] for i in range(len(row))}


def _upsert_bio(row):
    """Create/update candidate + profile from a dict of CANDIDATE_BIO fields."""
    email = (row.get("Email") or "").strip().lower()
    if not email:
        return None, False

    name = (row.get("Name") or "").strip()
    parts = name.split(maxsplit=1)
    first, last = (parts[0] if parts else email.split("@")[0]), (parts[1] if len(parts) > 1 else "")
    dob = _parse_date(row.get("DOB"))

    candidate = Candidate.objects.filter(email__iexact=email).first()
    created = False
    if not candidate:
        candidate = Candidate.objects.create(
            first_name=first,
            last_name=last,
            email=email,
            mobile=str(row.get("Mobile") or "").strip(),
            date_of_birth=dob,
        )
        created = True
    else:
        if first:
            candidate.first_name = first
        if last:
            candidate.last_name = last
        if dob:
            candidate.date_of_birth = dob
        mobile = str(row.get("Mobile") or "").strip()
        if mobile:
            candidate.mobile = mobile
        candidate.save()

    profile, _ = CandidateProfile.objects.get_or_create(candidate=candidate)
    gender = _norm(row.get("Gender"), _GENDER_MAP)
    if gender:
        profile.gender = gender
    category = _norm(row.get("Category"), _CATEGORY_MAP)
    if category:
        profile.category = category
    profile.is_pwbd = str(row.get("Is_PwBD") or "").strip().lower() in ("yes", "y", "true", "1")
    profile.aadhar_no = str(row.get("Aadhar_No") or "").strip()
    profile.permanent_address = str(
        row.get("Permanent_Address") or row.get("Permanent Address") or ""
    ).strip()
    profile.current_address = str(
        row.get("Current_Address") or row.get("Current Address") or ""
    ).strip()
    profile.current_same_as_permanent = not profile.current_address
    profile.save()
    return candidate, created


def import_bio_csv(fileobj, dry_run=False):
    """Import CANDIDATE_BIO rows from a CSV file."""
    reader = csv.DictReader(io.TextIOWrapper(fileobj, encoding="utf-8-sig"))
    stats = {"created": 0, "updated": 0, "skipped": 0, "rows": 0}
    for row in reader:
        stats["rows"] += 1
        candidate, created = _upsert_bio(row)
        if candidate is None:
            stats["skipped"] += 1
        elif dry_run:
            continue
        elif created:
            stats["created"] += 1
        else:
            stats["updated"] += 1
    return stats


def import_workbook(fileobj):
    """Import a TalentBridge .xlsx workbook (all four sheets)."""
    import openpyxl

    wb = openpyxl.load_workbook(fileobj, data_only=True)
    stats = {"created": 0, "updated": 0, "skipped": 0, "rows": 0, "academic": 0, "work": 0}

    candidate = None

    # --- CANDIDATE_BIO ---
    if "CANDIDATE_BIO" in wb.sheetnames:
        ws = wb["CANDIDATE_BIO"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1 and any(v is not None for v in rows[1]):
            header = rows[0]
            bio = _row_to_dict(header, rows[1])
            candidate, created = _upsert_bio(bio)
            if candidate is None:
                stats["skipped"] += 1
            else:
                stats["rows"] += 1
                stats["created" if created else "updated"] += 1
        else:
            stats["skipped"] += 1

    if candidate is None:
        return stats

    # --- ACADEMIC_RECORD ---
    if "ACADEMIC_RECORD" in wb.sheetnames:
        ws = wb["ACADEMIC_RECORD"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            for data_row in rows[1:]:
                if not any(v is not None and str(v).strip() for v in data_row):
                    continue
                r = _row_to_dict(header, data_row)
                level = _norm(r.get("Level"), _LEVEL_MAP)
                if not level:
                    level = "other"
                AcademicRecord.objects.create(
                    candidate=candidate,
                    level=level,
                    discipline=str(r.get("Discipline") or "").strip(),
                    university_board=str(r.get("University_Board") or "").strip(),
                    year_passed=_to_int(r.get("Year_Passed")),
                    marking_type=_norm(r.get("Marking_Type"), _MARKING_MAP, default="percentage"),
                    score=str(r.get("Score") or "").strip(),
                    is_ugc_recognized=str(r.get("Is_UGC_Recognized") or "").strip().lower()
                    in ("yes", "y", "true", "1"),
                )
                stats["academic"] += 1

    # --- WORK_EXP ---
    if "WORK_EXP" in wb.sheetnames:
        ws = wb["WORK_EXP"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            for data_row in rows[1:]:
                if not any(v is not None and str(v).strip() for v in data_row):
                    continue
                r = _row_to_dict(header, data_row)
                WorkExperience.objects.create(
                    candidate=candidate,
                    org_name=str(r.get("Org_Name") or "").strip() or "Unknown",
                    org_type=_norm(r.get("Org_Type"), {"psu": "psu", "private": "private", "govt": "govt"}),
                    designation=str(r.get("Designation") or "").strip(),
                    start_date=_parse_date(r.get("Start_Date")),
                    end_date=_parse_date(r.get("End_Date")),
                    annual_ctc_lakhs=_to_decimal(r.get("Annual_CTC_Lakhs")),
                    turnover_cr=_to_decimal(r.get("Turnover_Cr")),
                )
                stats["work"] += 1

    # --- EXAM_DISCLOSURE_PORTAL ---
    if "EXAM_DISCLOSURE_PORTAL" in wb.sheetnames:
        ws = wb["EXAM_DISCLOSURE_PORTAL"]
        rows = list(ws.iter_rows(values_only=True))
        if rows and any(v is not None and str(v).strip() for v in rows[1]):
            header = rows[0]
            r = _row_to_dict(header, rows[1])
            exam, _ = ExamDisclosure.objects.get_or_create(candidate=candidate)
            exam.exam_type = _norm(r.get("Exam_Type"), _EXAM_MAP)
            exam.gate_year = _to_int(r.get("GATE_Year"))
            exam.paper_code = str(r.get("Paper_Code") or "").strip()
            exam.marks_out_100 = _to_decimal(r.get("Marks_Out_100"))
            exam.gate_score = _to_decimal(r.get("GATE_Score"))
            exam.air = _to_int(r.get("AIR"))
            exam.ese_total_score = _to_decimal(r.get("ESE_Total_Score"))
            exam.public_disclosure_consent = str(r.get("Public_Disclosure_Consent") or "").strip().lower() in ("yes", "y", "true", "1")
            exam.save()

    return stats
